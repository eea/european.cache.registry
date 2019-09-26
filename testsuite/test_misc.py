import os
import json

from flask import url_for
from openpyxl import load_workbook

from instance.settings import FGAS, ODS, BDR_HELP_DESK_MAIL
from testsuite import factories
from cache_registry.models import MailAddress
from flask_mail import Mail

MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_export_companies(client):
    undertaking = factories.UndertakingFactory()
    resp = client.get(url_for('misc.export-company-list',
                              domain=undertaking.domain))
    assert resp.status_code == 200
    assert resp.content_type == MIMETYPE
    fn = 'test_file.xlsx'
    with open(fn, 'wb') as f:
        f.write(resp.body)
        f.close()
    wb = load_workbook(fn)
    assert len(wb.worksheets) == 1
    rows = [row for row in wb.worksheets[0].rows]
    assert len(rows) == 2
    assert rows[0][0].value == 'company_id'
    assert rows[1][0].value == undertaking.external_id
    assert rows[0][1].value == 'name'
    assert rows[1][1].value == undertaking.name
    assert rows[0][2].value == 'domain'
    assert rows[1][2].value == undertaking.domain
    assert rows[0][3].value == 'status'
    assert rows[1][3].value == undertaking.status
    assert rows[0][4].value == 'undertaking_type'
    assert rows[1][4].value == undertaking.undertaking_type
    assert rows[0][5].value == 'website'
    assert rows[1][5].value == undertaking.website
    os.remove(fn)


def test_export_companies_domain_filter(client):
    factories.UndertakingFactory(domain=ODS)
    undertaking = factories.UndertakingFactory(domain=FGAS)
    resp = client.get(url_for('misc.export-company-list',
                              domain=undertaking.domain))
    assert resp.status_code == 200
    assert resp.content_type == MIMETYPE
    fn = 'test_file.xlsx'
    with open(fn, 'wb') as f:
        f.write(resp.body)
        f.close()
    wb = load_workbook(fn)
    assert len(wb.worksheets) == 1
    rows = [row for row in wb.worksheets[0].rows]
    assert len(rows) == 2
    assert rows[0][0].value == 'company_id'
    assert rows[1][0].value == undertaking.external_id
    assert rows[0][1].value == 'name'
    assert rows[1][1].value == undertaking.name
    assert rows[0][2].value == 'domain'
    assert rows[1][2].value == undertaking.domain
    assert rows[0][3].value == 'status'
    assert rows[1][3].value == undertaking.status
    assert rows[0][4].value == 'undertaking_type'
    assert rows[1][4].value == undertaking.undertaking_type
    assert rows[0][5].value == 'website'
    assert rows[1][5].value == undertaking.website
    os.remove(fn)


def test_user_list_export(client):
    undertaking = factories.UndertakingFactory(domain=FGAS)
    user = factories.UserFactory(username='ExportUser',
                                 first_name='ExportFirstName',
                                 last_name='ExportLastName',
                                 email='ExportMail'
                                 )
    undertaking.contact_persons.append(user)
    resp = client.get(url_for('misc.export-user-list',
                              domain=FGAS))

    assert resp.status_code == 200
    assert resp.content_type == MIMETYPE
    fn = 'test_file.xlsx'
    with open(fn, 'wb') as f:
        f.write(resp.body)
        f.close()
    wb = load_workbook(fn)
    assert len(wb.worksheets) == 1
    rows =  [row for row in wb.worksheets[0].rows]
    assert rows[0][0].value == 'username'
    assert rows[1][0].value == user.username
    assert rows[0][1].value == 'companyname'
    assert rows[1][1].value == undertaking.name
    assert rows[0][2].value == 'country'
    assert rows[1][2].value == undertaking.address.country.name
    assert rows[0][3].value == 'contact_firstname'
    assert rows[1][3].value == user.first_name
    assert rows[0][4].value == 'contact_lastname'
    assert rows[1][4].value == user.last_name
    assert rows[0][5].value == 'contact_email'
    assert rows[1][5].value == user.email
    os.remove(fn)


def test_user_list_export_json(client):
    undertaking = factories.UndertakingFactory(domain=FGAS)
    user = factories.UserFactory(username='ExportUser',
                                 first_name='ExportFirstName',
                                 last_name='ExportLastName',
                                 email='ExportMail'
                                 )
    undertaking.contact_persons.append(user)
    resp = client.get(url_for('misc.export-user-list-json',
                              domain=FGAS))
    assert resp.status_code == 200
    data = resp.json
    assert data[0]['username'] == user.username
    assert data[0]['companyname'] == undertaking.name
    assert data[0]['country'] == undertaking.address.country.name
    assert data[0]['contact_firstname'] == user.first_name
    assert data[0]['contact_lastname'] == user.last_name
    assert data[0]['contact_email'] == user.email


def test_mail_address_list(client):
    ma = factories.MailAddress()
    resp = client.get(url_for('misc.mails-list'))
    assert resp.status_code == 200
    data = resp.json
    assert len(data) == 1
    assert data[0]['mail'] == ma.mail
    assert data[0]['first_name'] == ma.first_name
    assert data[0]['last_name'] == ma.last_name


def test_mail_address_add_new(client):
    resp = client.post(url_for('misc.mails-add'), dict(
        mail='a@ya.com', first_name='a', last_name='b'))
    assert resp.status_code == 200
    resp = json.loads(resp.body)
    assert resp['success'] is True
    assert len(MailAddress.query.all()) == 1
    ma = MailAddress.query.all()[0]
    assert ma.mail == 'a@ya.com'
    assert ma.first_name == 'a'
    assert ma.last_name == 'b'


def test_mail_address_add_existing(client):
    ma = factories.MailAddress()
    resp = client.post(url_for('misc.mails-add'), dict(
        mail=ma.mail, first_name='a', last_name='b'))
    assert resp.status_code == 200
    resp = json.loads(resp.body)
    assert resp['success'] is False
    assert len(MailAddress.query.all()) == 1


def test_mail_address_delete_existing(client):
    ma = factories.MailAddress()
    resp = client.post(url_for('misc.mails-delete'), dict(mail=ma.mail))
    assert resp.status_code == 200
    resp = json.loads(resp.body)
    assert resp['success'] is True
    assert len(MailAddress.query.all()) == 0


def test_mail_address_delete_nonexisting(client):
    factories.MailAddress()
    resp = client.post(url_for('misc.mails-delete'),
                       dict(mail='non@existing.com'))
    assert resp.status_code == 200
    resp = json.loads(resp.body)
    assert resp['success'] is False
    assert len(MailAddress.query.all()) == 1


def test_mail_alert_wrong_match(client):
    undertaking = factories.UndertakingFactory(oldcompany_verified=False)
    oldcompany = factories.OldCompanyFactory(id=2)
    factories.OldCompanyLinkFactory(oldcompany=oldcompany,
                                    undertaking=undertaking)
    user = factories.UserFactory()
    MailAddress.query.delete()
    resp = client.post(url_for('misc.alert-wrong-match',
                               domain=undertaking.domain),
                       dict(company_id=undertaking.external_id, user=user.id))
    assert resp.status_code == 200
    assert json.loads(resp.body)


def test_mail_alert_wrong_lockdown(client):
    undertaking = factories.UndertakingFactory(oldcompany_verified=False)
    oldcompany = factories.OldCompanyFactory(id=2)
    factories.OldCompanyLinkFactory(oldcompany=oldcompany,
                                    undertaking=undertaking)
    user = factories.UserFactory()
    factories.MailAddress()
    resp = client.post(url_for('misc.alert-wrong-lockdown',
                               domain=undertaking.domain),
                       dict(company_id=undertaking.external_id, user=user.id))
    assert resp.status_code == 200
    assert json.loads(resp.body)


def test_mail_alert_unmatch(client):
    undertaking = factories.UndertakingFactory(oldcompany_verified=False)
    oldcompany = factories.OldCompanyFactory()
    user = factories.UserFactory()
    factories.MailAddress()
    resp = client.post(
        url_for('misc.alert-unmatch', domain=undertaking.domain),
        dict(company_id=undertaking.external_id,
             old_company=oldcompany.external_id,
             user=user.id, oldcollection_path=''))
    assert resp.status_code == 200
    assert json.loads(resp.body)


def test_organization_log(client):
    factories.OrganizationLog(organizations=2)
    factories.MailAddress()
    resp = client.get(url_for('misc.log-sync', domain=FGAS))
    assert resp.status_code == 200
    data = json.loads(resp.body)
    assert len(data) == 1
    assert data[0]['domain'] == FGAS
    assert data[0]['organizations'] == 2
